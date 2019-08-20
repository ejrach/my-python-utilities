# This script uses OPENPYXL to read specific cells of data in an xlsx spreadsheet,
# perform some operation and write to different cells.
# This could be done to replace VB macros

# The "as xl" means it's just an alias so we can make our code shorter when using the library
import openpyxl as xl

# Just open two classes
from openpyxl.chart import BarChart, Reference

# Define constants here
#EXCEL_INPUT = "input.xlsx"
#EXCEL_OUTPUT = "output.xlsx"
STARTING_ROW = 2  # we don't want to include the header row of the file.
COLUMN_TO_UPDATE = 3
COLUMN_CORRECTED_DATA = 4


# def process_workbook():
def process_workbook(filename):
    # Function to process a workbook

    # First make a copy of the input file, give it a new name
    # and then open it for modifications. We don't want to disturb master copy
    try:

        # Load the workbook
        wb = xl.load_workbook(filename)

        # Specify the sheet name, because we are going to access specific cells
        sheet = wb['Sheet1']

        # We can access a cell by it's column letter and row number
        cell = sheet['a1']
        # Or we can access by column number and row number
        cell = sheet.cell(1, 1)

        # Loop from the 2nd row (don't include the header) to the last row.
        # Adding 1 because of the header
        for row in range(STARTING_ROW, sheet.max_row + 1):
            cell = sheet.cell(row, COLUMN_TO_UPDATE)
            corrected_price = cell.value * 0.9

            # Where to put the modified data, i.e. in the same row, but in the
            # specified column
            corrected_price_cell = sheet.cell(row, COLUMN_CORRECTED_DATA)
            corrected_price_cell.value = corrected_price

        # Now create the chart.
        # min_row is telling us which row to start from (i.e. where the data starts)
        # max_row is telling us to go to the end of the data rows
        # min_col and max_col is indicating which data columns to display
        # values = Reference(sheet,
        #                    min_row=2,
        #                    max_row=sheet.max_row,
        #                    min_col=4,
        #                    max_col=4)

        # chart = BarChart()
        # chart.add_data(values)
        # NOTE: This line doesn't work in python3. Tells us where to add the corner of the upper left hand corner of the chart
        # sheet.add_chart(chart, 'e2')

        wb.save(filename)

    except:
        print("Unable to process file")


# Process a workbook
process_workbook("input.xlsx")
# process_workbook()
